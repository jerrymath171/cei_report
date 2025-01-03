#  python
# -*- coding: utf-8 -*- 
# @Time : 2024/11/29 13:39
# @Author : Mth@cei 
# @File : error_finding2.0.py  
# @Software: PyCharm

import string
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import os
from tqdm import tqdm


def generate_excel_column_names(num_columns):
    """
    Generate a list of Excel column names up to a given number.

    This function creates Excel-style column names (e.g., A, B, ..., Z, AA, AB, ..., AZ, BA, ...)
    for a specified number of columns. It uses a cycle of letters to generate column names
    in the same manner as Excel.

    Args:
    num_columns (int): The number of Excel column names to generate.

    Returns:
    list: A list of strings representing Excel column names up to the specified number.
    """
    names = []
    letters = string.ascii_uppercase
    for i in range(1, num_columns + 1):
        name = ''
        n = i
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            name = letters[remainder] + name
        names.append(name)
    return names


def str_to_date(value):
    return pd.to_datetime(datetime.datetime.strptime(str(value), '%Y%m'))


def usc_count(value, header='91'):
    if pd.isnull(value):
        return 0
    else:
        if len(value) != 18:
            return 0
        if header is None:
            return 1
        else:
            if str(value).startswith(header):
                return 1
            else:
                return 0


def isnull_count(value):
    if pd.isnull(value):
        return 0
    else:
        if str(value) in ['无', '/', '//', 'N/A', '-', r'\\']:
            return 0
        else:
            return 1


def deposit_count(value, date, content):
    try:
        if int(value) == 0:
            return 0, []
    except ValueError:
        pass
    if pd.isnull(value):
        return 0, []
    else:
        correct_flag = 1
        value_list = value.split('；')
        correct_value = value.replace(';', '；')
        correct_value_list = correct_value.split('；')
        # 检查分号是否正确，若不正确，用正确的替换
        if value_list != correct_value_list:
            correct_flag = 0
        value_list = correct_value_list
        deposit_dict = dict()
        effective_count = 0
        for item in value_list:
            if len(item) == 0:
                continue
            effective_count += 1
            item_list = item.split('，')
            correct_item = item.replace(',', '，')
            correct_item_list = correct_item.split('，')
            # 检查逗号是否正确，若不正确，用正确的替换
            if item_list != correct_item_list:
                correct_flag = 0
                item_list = correct_item_list
            if str(item_list[1]) == '0':
                correct_flag = 0
            elif '.' in str(item_list[1]) and content == 'person':
                correct_flag = 0
            if month_count(item_list[0]) == 0 or isdigit_count(item_list[1]) == 0:
                correct_flag = 0
            # 记录正确的数据字典
            deposit_dict[str_to_date(item_list[0])] = item_list[1]
        if correct_flag == 0:
            return 0, sorted(deposit_dict.keys())
        if len(deposit_dict.keys()) != effective_count:
            return 0, sorted(deposit_dict.keys())
        else:
            pass
        if max(deposit_dict.keys()) != date:
            return 0, sorted(deposit_dict.keys())
        else:
            pass
        if min(deposit_dict.keys()) < date - pd.DateOffset(years=2):
            return 0, sorted(deposit_dict.keys())
        else:
            return 1, sorted(deposit_dict.keys())


def water_bill_count(value):
    """Check if a given water bill record is valid.

    Args:
    value (str): The water bill record to check.

    Returns:
    int: 1 if the record is valid, 0 otherwise.
    """
    if pd.isnull(value):
        return 1

    records = value.split('；')
    for record in records:
        if len(record) == 0:
            continue

        items = record.split('，')
        if len(items) != 2:
            return 0

        date, amount = items
        if amount == '0':
            return 0

        if not isdigit_count(amount):
            return 0

        if not is_date_count(date):
            return 0

    return 1


def continuous_month_count(value, person_list, amount_list):
    if person_list != amount_list:
        return 0
    if pd.isnull(value):
        return 0
    else:
        try:
            # 基础检验无误，开始检验数量是否正确
            if int(value) == 0:
                # 值为0，且也查不到记录，那么正确
                if len(person_list) == 0:
                    return 1
                else:
                    return 0
            elif 0 < int(value) <= 24:
                # 若记录数比连续数少，那么肯定错误
                if len(person_list) < int(value):
                    return 0
                continuous_count = 1
                max_count = 1
                # 计算连续性
                for i in range(1, len(person_list)):
                    if person_list[i - 1] + pd.DateOffset(months=1) == person_list[i]:
                        continuous_count += 1
                        if continuous_count > max_count:
                            max_count = continuous_count
                    else:
                        continuous_count = 1
                if max_count == int(value):
                    return 1
                else:
                    return 0
            else:
                return 0
        except ValueError:
            return 0


def last_date_count(value, person_list, amount_list):
    if person_list != amount_list:
        return 0
    if len(person_list) == 0:
        return 1
    if pd.isnull(value):
        return 0
    else:
        if len(str(value)) != 6:
            return 0
        else:
            try:
                # 检查日期是否为最后一个
                if str_to_date(value) <= check_date and str_to_date(value) == person_list[-1]:
                    return 1
                else:
                    return 0
            except ValueError:
                return 0


def zero_one_count(value):
    if pd.isnull(value):
        return 0
    else:
        try:
            if int(value) == 1 or int(value) == 0:
                return 1
            else:
                return 0
        except ValueError:
            return 0


def isdigit_count(value, optional=0, options=None):
    if pd.isnull(value):
        return optional
    else:
        try:
            if options is not None and float(value) in options:
                return 1
            if float(value) > 0:
                return 1
            else:
                return 0
        except ValueError:
            return 0


def correct_month_count(value, person_list, amount_list):
    if person_list != amount_list:
        return 0
    # 如果查不到记录，默认是对的
    if len(person_list) == 0:
        return 1
    if pd.isnull(value):
        return 0
    else:
        try:
            if 0 <= int(value) <= 24 and int(value) == len(person_list):
                return 1
            else:
                return 0
        except ValueError:
            return 0


def is_date_count(value, optional=0):
    """
    Check if a given value is a valid date.

    Args:
    value (str): The value to check.

    Returns:
    int: 1 if the value is a valid date, 0 otherwise.
    """
    if pd.isnull(value):
        return optional

    # List of date formats to check
    date_formats = ['%Y%m%d', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']

    # Iterate over the date formats
    for date_format in date_formats:
        try:
            datetime.datetime.strptime(str(value), date_format)
            return 1
        except ValueError:
            continue

    # If no valid date format was found, return 0
    return 0


def month_count(value, optional=0):
    """
    Check if a given value is a valid date.

    Args:
    value (str): The value to check.

    Returns:
    int: 1 if the value is a valid date, 0 otherwise.
    """
    if pd.isnull(value):
        return optional

    # List of date formats to check
    date_formats = ['%Y%m', '%Y-%m', '%Y/%m', '%Y.%m']

    # Iterate over the date formats
    for date_format in date_formats:
        try:
            datetime.datetime.strptime(str(value), date_format)
            return 1
        except ValueError:
            continue

    # If no valid date format was found, return 0
    return 0


def time_count(value, optional=0):
    """
    Check if a given value is a valid time.

    Args:
    value (str): The value to check.

    Returns:
    int: 1 if the value is a valid time, 0 otherwise.
    """
    if pd.isnull(value):
        return optional

    # List of date formats to check
    date_formats = ['%Y%m%d%H%M%S', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y.%m.%d %H:%M:%S']

    # Iterate over the date formats
    for date_format in date_formats:
        try:
            datetime.datetime.strptime(str(value), date_format)
            return 1
        except ValueError:
            continue

    # If no valid date format was found, return 0
    return 0


def yesno_count(value, optional=0):
    if pd.isnull(value):
        return optional
    else:
        try:
            if value == '是' or value == '否':
                return 1
            else:
                return 0
        except ValueError:
            return 0


def period_count(value):
    if pd.isnull(value):
        return 0
    else:
        time_list = value.split('-')
        if len(time_list) != 2:
            return 0
        else:
            if is_date_count(time_list[0]) and is_date_count(time_list[1]):
                return 1
            else:
                return 0


def base_date_count(value, date):
    if pd.isnull(value):
        return 0
    else:
        # List of date formats to check
        date_formats = ['%Y%m%d', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']
        date_value = None

        # Iterate over the date formats
        for date_format in date_formats:
            try:
                date_value = datetime.datetime.strptime(str(value), date_format)
                return 1
            except ValueError:
                continue
        if pd.to_datetime(date_value) == date + pd.DateOffset(months=1) - pd.DateOffset(days=1):
            return 1
        else:
            return 0


def two_years_count(value):
    if pd.isnull(value):
        return 0
    else:
        try:
            if 0 <= int(value) <= 24:
                return 1
            else:
                return 0
        except ValueError:
            return 0


def ele_status_count(value, options, optional=0):
    if pd.isnull(value):
        return optional
    else:
        if str(value) in options:
            return 1
        else:
            if str(value).startswith('其他-'):
                return 1
            else:
                return 0


def identification_count(value, optional=0):
    if pd.isnull(value):
        return optional
    else:
        if len(value) <= 20:
            return 0
        else:
            return 1


def range_count(value, upper, options=None):
    if options is not None and str(value) in options:
        return 1
    if pd.isnull(value):
        return 0
    else:
        try:
            if 1 <= int(value) <= upper:
                return 1
            else:
                return 0
        except ValueError:
            return 0


def integer_count(value, options=None):
    try:
        if options is not None and int(value) in options:
            return 1
        else:
            pass
    except ValueError:
        pass
    if pd.isnull(value):
        return 0
    else:
        try:
            if int(value) > 0:
                return 1
            else:
                return 0
        except ValueError:
            return 0


def compare_count(value1, value2, function):
    if pd.isnull(value1) or pd.isnull(value2):
        return 0
    else:
        if function(value1) == 0 or function(value2) == 0:
            return 0
        try:
            if float(value1) >= float(value2):
                return 1
            else:
                return 0
        except ValueError:
            return 0


def insurance_count(value, date):
    if pd.isnull(value):
        return 0
    else:
        value_list = value.split('；')
        month_list = []
        for item in value_list:
            if len(item) == 0:
                continue
            item_list = item.split('，')
            if len(item_list) != 4:
                return 0
            else:
                if is_date_count(item_list[0]) == 0:
                    return 0
                try:
                    int(item_list[1])
                except ValueError:
                    return 0
                if '.' in item_list[1]:
                    return 0
                if isdigit_count(item_list[2], options=[0]) == 0:
                    return 0
                if month_count(item_list[3]) == 0:
                    return 0
                month_list.append(pd.to_datetime(item_list[0]))
        if max(month_list) < date or max(month_list) > date + pd.DateOffset(months=1):
            return 0
        if min(month_list) < date - pd.DateOffset(years=2):
            return 0
        return 1


def delay_count(value):
    if pd.isnull(value):
        return 0
    if str(value) == '无':
        return 1
    value_list = value.split('；')
    for item in value_list:
        if len(item) == 0:
            continue
        item_list = item.split('，')
        if len(item_list) != 3:
            return 0
        if is_date_count(item_list[0]) == 0:
            return 0
        if isdigit_count(item_list[1]):
            return 0
        if isdigit_count(item_list[2]) == 0:
            return 0
    return 1


def rate_count(value, date):
    if pd.isnull(value):
        return 0
    contents_list = value.split('。')
    contents_length = 0
    temp_type = None
    for contents in contents_list:
        if len(contents) == 0:
            continue
        contents_length += 1
        item_list = contents.split('；')
        item_length = 0
        for i, item in enumerate(item_list):
            if len(item) == 0:
                continue
            item_length += 1
            temp_list = item.split('，')
            if len(temp_list) != 3:
                return 0
            if month_count(temp_list[0]) == 0:
                return 0
            if str_to_date(temp_list[0]) != (date - pd.DateOffset(months=i)):
                return 0
            if i == 0:
                if temp_type == temp_list[1]:
                    return 0
                temp_type = temp_list[1]
                if temp_type not in ['同比', '环比']:
                    return 0
            else:
                if temp_type != temp_list[1]:
                    return 0
            if '%' not in temp_list[2]:
                return 0
            try:
                float(temp_list[2].split('%')[0])
            except ValueError:
                return 0
            if len(str(temp_list[2]).split('.')) > 2:
                return 0
        if item_length != 3:
            return 0
    if contents_length != 2:
        return 0
    return 1


def house_check(file, date):
    color_result = pd.DataFrame(index=file['企业统一社会信用代码'], columns=file.columns)
    temp_person, temp_amount = [], []
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '企业统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '近两年每月缴存人数':
                color_result.iloc[i, j], temp_person = deposit_count(file.iloc[i, j], date, 'person')
            elif file.columns[j] == '近两年每月当月住房公积金缴存额':
                color_result.iloc[i, j], temp_amount = deposit_count(file.iloc[i, j], date, 'deposit')
            elif file.columns[j] == '近两年连续正常缴纳周期（月））':
                color_result.iloc[i, j] = continuous_month_count(file.iloc[i, j], temp_person, temp_amount)
            elif file.columns[j] == '最近一次正常缴费月份':
                color_result.iloc[i, j] = last_date_count(file.iloc[i, j], temp_person, temp_amount)
            elif file.columns[j] == '当前是否欠费':
                color_result.iloc[i, j] = zero_one_count(file.iloc[i, j])
            elif file.columns[j] == '欠缴金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '近两年单位正常缴费次数':
                color_result.iloc[i, j] = correct_month_count(file.iloc[i, j], temp_person, temp_amount)
    return color_result


def insurance_check(file, date):
    color_result = pd.DataFrame(index=file['企业统一社会信用代码'], columns=file.columns)
    temp_person, temp_amount = [], []
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业职工基本养老保险近两年每月参保人数':
                color_result.iloc[i, j], temp_person = deposit_count(file.iloc[i, j], date, 'person')
            elif file.columns[j] == '企业职工基本养老保险近两年每月单位缴费金额':
                color_result.iloc[i, j], temp_amount = deposit_count(file.iloc[i, j], date, 'deposit')
            elif file.columns[j] == '企业职工基本养老保险近两年连续正常缴纳周期（月）':
                color_result.iloc[i, j] = continuous_month_count(file.iloc[i, j], temp_person, temp_amount)
            elif file.columns[j] == '企业职工基本养老保险最近一次正常缴费月份':
                color_result.iloc[i, j] = last_date_count(file.iloc[i, j], temp_person, temp_amount)
            elif file.columns[j] == '企业职工基本养老保险当前是否欠缴':
                color_result.iloc[i, j] = zero_one_count(file.iloc[i, j])
            elif file.columns[j] == '企业职工基本养老保险欠缴金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '企业职工基本养老保险近两年单位正常缴费次数':
                color_result.iloc[i, j] = correct_month_count(file.iloc[i, j], temp_person, temp_amount)
    return color_result


def water_check(file, date):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '水用户信息(户名户号)':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '开户日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '用户地址':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '预交金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '欠费金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '是否一户一表':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j])
            elif file.columns[j] == '缴纳水费明细':
                color_result.iloc[i, j] = water_bill_count(file.iloc[i, j])
            elif file.columns[j] == '近3个月月均用水量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1, options=[0])
            elif file.columns[j] == '近6个月月均用水量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1, options=[0])
            elif file.columns[j] == '当前是否欠费':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j], 1)
            elif file.columns[j] == '缴费所属期':
                color_result.iloc[i, j] = period_count(file.iloc[i, j])
            elif file.columns[j] == '统计日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '统计基准日期':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
    return color_result


def electricity_check(file, date):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '电网户号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '用户地址':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '运行容量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '合同容量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '首次供电时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '用电账户状态':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j],
                                                           ['正常', '欠费', '预付费', '暂停供电', '销户异常'])
            elif file.columns[j] == '欠费金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '近两年违约次数':
                color_result.iloc[i, j] = two_years_count(file.iloc[i, j])
            elif file.columns[j] == '用电类型':
                color_result.iloc[i, j] = ele_status_count(
                    file.iloc[i, j], ['居民生活用电', '一般工商业用电', '大工业用电', '农业生产用电'], 1)
            elif file.columns[j] == '价值等级':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], ['基础用电', '正常用电', '高耗能用电'], 1)
            elif file.columns[j] == '风险等级':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], ['低风险', '中风险', '高风险', '极高风险'],
                                                           1)
            elif file.columns[j] == '近3个月月均用电金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1)
            elif file.columns[j] == '近6个月月均用电金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1)
            elif file.columns[j] == '近一年月均用电金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1)
            elif file.columns[j] == '当前是否欠费':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j], 1)
            elif file.columns[j] == '统计日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '统计基准日期':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
    return color_result


def gas_check(file, date):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '燃气用户信息(户名户号)':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '开户日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '用户地址':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '预交金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '欠费金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '是否一户一表':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j])
            elif file.columns[j] == '缴纳燃气费明细':
                color_result.iloc[i, j] = water_bill_count(file.iloc[i, j])
            elif file.columns[j] == '缴费所属期':
                color_result.iloc[i, j] = period_count(file.iloc[i, j])
            elif file.columns[j] == '近3个月月均用燃气量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1)
            elif file.columns[j] == '近6个月月均用燃气量':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], 1)
            elif file.columns[j] == '当前是否欠费':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j], 1)
            elif file.columns[j] == '统计日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '统计基准日期':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
    return color_result


def enterprise_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
                # TODO: isnull_count会返回0或1，当返回1时，请在这里加一个验证此企业是否存续的判断，如果存续返回1，否则返回0，传给color_result.iloc[i, j]
                # color_result.iloc[i, j] = 0或1
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j], header='91')
            elif file.columns[j] == '所属行业':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '登记机关':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '注册资本':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '注册资本币种':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '核准日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '参保人数':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '企业曾用名':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '经营范围':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def enterprise_report_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '地方平台编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '是否实名认证':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j], 1)
            elif file.columns[j] == '实名认证人身份':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], ['法人', '董监高', '财务负责人', '其他'], 1)
            elif file.columns[j] == '实名认证人身份证号':
                color_result.iloc[i, j] = identification_count(file.iloc[i, j], 1)
            elif file.columns[j] == '是否通过平台获得贷款':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j], 1)
            elif file.columns[j] == '企业所属行业':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 19)
            elif file.columns[j] == '企业所在省':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业所在市':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '注册资本':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '经营范围':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '经营期限类型':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 2)
            elif file.columns[j] == '营业期限开始日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j], 1)
            elif file.columns[j] == '营业期限结束日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j], 1)
            elif file.columns[j] == '核准日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j], 1)
            elif file.columns[j] == '入驻时间':
                color_result.iloc[i, j] = time_count(file.iloc[i, j])
            elif file.columns[j] == '外部系统id':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def financial_check(file):
    color_result = pd.DataFrame(index=file['机构统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '地方平台编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '机构全称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '机构统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '金融机构法人编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '机构类型':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 11, ['99'])
            elif file.columns[j] == '简介':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '入驻地方平台时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '金融产品数':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '金融机构所在省':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '外部系统id':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def loan_check(file):
    color_result = pd.DataFrame(index=file['获贷企业统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '地方平台编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '获贷企业统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j], header='91')
            elif file.columns[j] == '获贷企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '是否首贷':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j])
            elif file.columns[j] == '获贷时间':
                color_result.iloc[i, j] = time_count(file.iloc[i, j])
            elif file.columns[j] == '放款金融机构名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '放贷金融机构统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '放款金融机构法人编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '获贷金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '贷款利率':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '担保方式':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 7, ['0', '99'])
            elif file.columns[j] == '贷款期限':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '还款方式':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 6, ['0', '99'])
            elif file.columns[j] == '贷款状态':
                color_result.iloc[i, j] = range_count(file.iloc[i, j], 3)
    return color_result


def private_enterprise_check(file, date):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '平台编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业注册数量':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '授信民营企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '累计获贷民营企业数量':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业授信总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业累计获贷金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业累计信用贷款获贷金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业信用放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业融资需求金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业融资需求笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '民营企业放款平均利率':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '当月民营企业获贷数量':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业获贷金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业信用贷款放贷金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业信用放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业融资需求金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月民营企业融资需求笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '统计开始时间':
                color_result.iloc[i, j] = time_count(file.iloc[i, j])
            elif file.columns[j] == '统计截止时间(月)':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
            elif file.columns[j] == '外部系统id':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def capitalize_platform_check(file, date):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '平台编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '注册企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '授信企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '获贷企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '获贷小微企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], [0])
            elif file.columns[j] == '授信总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '放款总额':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['信用放款总额'].iloc[i], isdigit_count)
            elif file.columns[j] == '普惠小微贷款总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '信用放款总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '信用放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '融资需求金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '融资需求笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '放款平均利率':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '入驻金融机构数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '发布金融产品数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月放款金额':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['当月信用放款金额'].iloc[i],
                                                        isdigit_count)
            elif file.columns[j] == '当月放款笔数':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['当月信用放款笔数'].iloc[i],
                                                        integer_count)
            elif file.columns[j] == '当月新增普惠小微贷款金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月融资需求笔数':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '统计开始时间':
                color_result.iloc[i, j] = time_count(file.iloc[i, j])
            elif file.columns[j] == '统计截止时间(月)':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
            elif file.columns[j] == '外部系统id':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result, file[['授信总额', '信用放款总额', '放款总额', '普惠小微贷款总额']].sum()


def capitalize_institution_check(file, date):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '省/直辖市/自治区名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '行政区划代码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '金融机构全称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '银行业金融机构法人编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '授信企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '获贷企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '获贷小微企业数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '授信总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '放款总额':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['信用放款总额'].iloc[i], isdigit_count)
            elif file.columns[j] == '普惠小微贷款总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '信用放款总额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '信用放款笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '融资需求金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '融资需求笔数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '放款平均利率':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '入驻金融机构数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j])
            elif file.columns[j] == '发布金融产品数':
                color_result.iloc[i, j] = integer_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月放款金额':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['当月信用放款金额'].iloc[i],
                                                        isdigit_count)
            elif file.columns[j] == '当月放款笔数':
                color_result.iloc[i, j] = compare_count(file.iloc[i, j], file['当月信用放款笔数'].iloc[i],
                                                        integer_count)
            elif file.columns[j] == '当月新增普惠小微贷款金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '当月融资需求笔数':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '统计开始时间':
                color_result.iloc[i, j] = time_count(file.iloc[i, j])
            elif file.columns[j] == '统计截止时间(月)':
                color_result.iloc[i, j] = base_date_count(file.iloc[i, j], date)
            elif file.columns[j] == '外部系统id':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result, file[['授信总额', '信用放款总额', '放款总额', '普惠小微贷款总额']].sum()


def security_check(file, date):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '近两年每月参保人数、缴费金额及费款所属期':
                color_result.iloc[i, j] = insurance_count(file.iloc[i, j], date)
            elif file.columns[j] == '最近一次正常缴费月份':
                color_result.iloc[i, j] = month_count(file.iloc[i, j])
            elif file.columns[j] == '近两年每月缓缴金额':
                color_result.iloc[i, j] = delay_count(file.iloc[i, j])
    return color_result


def debtor_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '立案时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '案号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '执行法院':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '执行依据文号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def judgment_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '案号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '案件类型':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '文书类型':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '判决法院':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '裁判日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '案由':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '文书标题':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '公布日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '当事人诉讼地位':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '文书内容':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '案件金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[-1])
    return color_result


def qualification_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '资质证书名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '资质证书编号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '资质或证书类别及等级':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '有效期至':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '发证机关':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '发证日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
    return color_result


def honor_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '荣誉表彰名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '荣誉表彰事项及等级':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '颁发机关':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '颁发日期':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
    return color_result


def insurance_unpaid_check(file):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '近两年欠缴金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j], options=[0])
            elif file.columns[j] == '欠缴所属期':
                color_result.iloc[i, j] = period_count(file.iloc[i, j])
    return color_result


def insurance_change_check(file, date):
    color_result = pd.DataFrame(index=file['统一社会信用代码'], columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '企业名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统一社会信用代码':
                color_result.iloc[i, j] = usc_count(file.iloc[i, j])
            elif file.columns[j] == '近一季度单位月度参保人数变化率':
                color_result.iloc[i, j] = rate_count(file.iloc[i, j], date)
            elif file.columns[j] == '近一季度单位月度缴费金额变化率':
                color_result.iloc[i, j] = rate_count(file.iloc[i, j], date)
            elif file.columns[j] == '统计时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
    return color_result


def immovable_registry_check(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '权利人名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '权利人证件号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '不动产权证书号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '不动产单元号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '登记时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
            elif file.columns[j] == '用途':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '坐落':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '面积':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '使用期限':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '登记机构':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def immovable_mortgage_check(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '抵押权人':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '抵押权人证件号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '不动产登记证明号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '不动产单元号':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '权利类型':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '抵押人':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '抵押登记时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
    return color_result


def agricultural_insurance_check(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '新型农业经营主体名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '证件类型':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['统一社会信用代码', '身份证'])
            elif file.columns[j] == '证件号码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '是否投保':
                color_result.iloc[i, j] = yesno_count(file.iloc[i, j])
            elif file.columns[j] == '投保类型':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['农业保险', '财产保险', '责任保险', '信用保险'])
            elif file.columns[j] == '投保规模':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '保险金额':
                color_result.iloc[i, j] = isdigit_count(file.iloc[i, j])
            elif file.columns[j] == '保单有效期自':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '保单有效期至':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '保险理赔信息':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def agricultural_subsidies_check(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '新型农业经营主体名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '证件类型':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['统一社会信用代码', '身份证'])
            elif file.columns[j] == '证件号码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '补贴类型':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '补贴金额':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '统计时间':
                color_result.iloc[i, j] = is_date_count(file.iloc[i, j])
    return color_result


def large_growers_check(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '种植大户名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '证件类型':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['统一社会信用代码', '身份证'])
            elif file.columns[j] == '证件号码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def agricultural_entities(file):
    color_result = pd.DataFrame(index=file.index, columns=file.columns)
    for i in tqdm(range(file.shape[0])):
        for j in range(file.shape[1]):
            if file.columns[j] == '数据编码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '新型农业经营主体名称':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '负责人姓名':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '证件类型':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['统一社会信用代码', '身份证'])
            elif file.columns[j] == '证件号码':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '类别':
                color_result.iloc[i, j] = ele_status_count(file.iloc[i, j], options=['家庭农场', '龙头企业', '专业合作社', '农业大户'])
            elif file.columns[j] == '所在地':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '经营内容':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '名下所有经营地块信息':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '种养殖类型':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
            elif file.columns[j] == '种养殖规模':
                color_result.iloc[i, j] = isnull_count(file.iloc[i, j])
    return color_result


def main_check(name_type, name, check_type, filled_color, date):
    file = pd.DataFrame()
    if name_type == '1':
        name_list = name.split(' ')
        for temp_file in name_list:
            file = pd.concat([file, pd.read_excel(temp_file)], axis=0, ignore_index=True)
    elif name_type == '2':
        for temp_file in os.listdir(name):
            file = pd.concat([file, pd.read_excel(name + '\\' + temp_file)], axis=0, ignore_index=True)
    else:
        pass
    file = file.astype(str)
    charter_list = generate_excel_column_names(file.shape[1])
    if check_type == '1':
        color_result = house_check(file, date)
    elif check_type == '2':
        color_result = insurance_check(file, date)
    elif check_type == '3':
        color_result = water_check(file, date)
    elif check_type == '4':
        color_result = electricity_check(file, date)
    elif check_type == '5':
        color_result = gas_check(file, date)
    elif check_type == '6':
        color_result = enterprise_check(file)
    elif check_type == '7':
        color_result = enterprise_report_check(file)
    elif check_type == '9':
        color_result = financial_check(file)
    elif check_type == '10':
        color_result = loan_check(file)
    elif check_type == '11':
        color_result = private_enterprise_check(file, date)
    elif check_type == '12':
        color_result = security_check(file, date)
    elif check_type == '13':
        color_result = debtor_check(file)
    elif check_type == '14':
        color_result = judgment_check(file)
    elif check_type == '15':
        color_result = qualification_check(file)
    elif check_type == '16':
        color_result = honor_check(file)
    elif check_type == '17':
        color_result = insurance_unpaid_check(file)
    elif check_type == '18':
        color_result = insurance_change_check(file, date)
    elif check_type == '19':
        color_result = agricultural_insurance_check(file)
    elif check_type == '20':
        color_result = agricultural_subsidies_check(file)
    elif check_type == '21':
        color_result = large_growers_check(file)
    elif check_type == '22':
        color_result = agricultural_entities(file)
    else:
        color_result = pd.DataFrame()
        print('输入错误')
    color_result = color_result.fillna(1)
    print('正在输出文件')
    correct_id = set(color_result[color_result.sum(axis=1) == color_result.shape[1]].index)
    correct_num = len(correct_id)
    problem_sheet = pd.DataFrame()
    problem_list = ['合规总数：{}'.format(str(correct_num))]
    problem_sheet['问题汇总'] = problem_list
    if name_type == '1':
        writer_name = name.split(os.sep)[-1].split('.')[0] + '合规记录统计.xlsx'
        writer = pd.ExcelWriter(writer_name)
        file.to_excel(writer, index=False, sheet_name='Sheet1')
        problem_sheet.to_excel(writer, index=False, sheet_name='Sheet2')
        writer.close()
    elif name_type == '2':
        writer_name = name.split(os.sep)[-1] + '合规记录统计.xlsx'
        writer = pd.ExcelWriter(writer_name)
        file.to_excel(writer, index=False, sheet_name='Sheet1')
        problem_sheet.to_excel(writer, index=False, sheet_name='Sheet2')
        writer.close()
    else:
        writer_name = ''
        pass
    print('正在对文件上色')
    wb = load_workbook(filename=writer_name)
    work = wb[wb.sheetnames[0]]
    for m in range(color_result.shape[0]):
        for n in range(color_result.shape[1]):
            if color_result.iloc[m, n] == 0:
                work[charter_list[n] + str(m + 2)].fill = filled_color
    wb.close()
    wb.save(writer_name)


def double_check(name_type, name, filled_color, date):
    file_finance, file_platform = pd.DataFrame(), pd.DataFrame()
    finance_name, platform_name = '', ''
    if name_type == '1':
        name_list = name.split(' ')
        for temp_file in name_list:
            if '金融机构' in temp_file and '融资统计' in temp_file:
                finance_name = temp_file
                file_finance = pd.read_excel(temp_file)
                break
        for temp_file in name_list:
            if '平台' in temp_file and '融资统计' in temp_file:
                platform_name = temp_file
                file_platform = pd.read_excel(temp_file)
                break
    elif name_type == '2':
        for temp_file in os.listdir(name):
            if temp_file.startswith('~$'):
                continue
            if '金融机构' in temp_file and '融资统计' in temp_file:
                finance_name = temp_file
                file_finance = pd.read_excel(name + '\\' + temp_file)
                break
        for temp_file in os.listdir(name):
            if temp_file.startswith('~$'):
                continue
            if '平台' in temp_file and '融资统计' in temp_file:
                platform_name = temp_file
                file_platform = pd.read_excel(name + '\\' + temp_file)
                break
    else:
        pass
    color_institution, finance_sum = capitalize_institution_check(file_finance, date)
    color_platform, platform_sum = capitalize_platform_check(file_platform, date)
    conflict_list = []
    for index_name in finance_sum.index:
        if finance_sum[index_name] == platform_sum[index_name]:
            conflict_list.append(index_name)

    print('正在输出金融机构表')
    # 输出金融机构表
    color_institution_id = set(color_institution[color_institution.sum(axis=1) == color_institution.shape[1]].index)
    color_institution_num = len(color_institution_id)
    problem_sheet = pd.DataFrame()
    problem_list = ['合规总数：{}'.format(str(color_institution_num))]
    problem_sheet['问题汇总'] = problem_list
    writer_name = finance_name.split('.')[0] + '合规记录统计.xlsx'
    writer_finance = pd.ExcelWriter(writer_name)
    file_finance.to_excel(writer_finance, index=False, sheet_name='Sheet1')
    problem_sheet.to_excel(writer_finance, index=False, sheet_name='Sheet2')
    writer_finance.close()
    wb = load_workbook(filename=writer_name)
    work = wb[wb.sheetnames[0]]
    charter_list = generate_excel_column_names(file_finance.shape[1])
    for m in range(color_institution.shape[0]):
        for n in range(color_institution.shape[1]):
            if color_institution.iloc[m, n] == 0:
                work[charter_list[n] + str(m + 2)].fill = filled_color
    wb.close()
    wb.save(writer_name)

    print('正在输出平台表')
    # 输出平台表
    color_platform_id = set(color_institution[color_institution.sum(axis=1) == color_institution.shape[1]].index)
    color_platform_num = len(color_platform_id)
    problem_sheet = pd.DataFrame()
    problem_list = ['合规总数：{}'.format(str(color_platform_num))]
    problem_sheet['问题汇总'] = problem_list
    writer_name = platform_name.split('.')[0] + '合规记录统计.xlsx'
    writer_platform = pd.ExcelWriter(writer_name)
    file_platform.to_excel(writer_platform, index=False, sheet_name='Sheet1')
    problem_sheet.to_excel(writer_platform, index=False, sheet_name='Sheet2')
    writer_platform.close()
    wb = load_workbook(filename=writer_name)
    work = wb[wb.sheetnames[0]]
    charter_list = generate_excel_column_names(file_platform.shape[1])
    for m in range(color_platform.shape[0]):
        for n in range(color_platform.shape[1]):
            if color_platform.iloc[m, n] == 0:
                work[charter_list[n] + str(m + 2)].fill = filled_color
    wb.close()
    wb.save(writer_name)


if __name__ == '__main__':
    while True:
        fill = PatternFill("solid", fgColor='FFFF00')
        input_date = input('请输入检查日期（格式：YYYYMM)')
        check_date = str_to_date(input_date)
        file_type = input('请选择输入文件名或文件夹名。1：文件名 2：文件夹名')
        if file_type == '1':
            file_name = input('请输入要检查的文件名（多个文件名以空格分隔）')
        elif file_type == '2':
            file_name = input('请将要检查的同一类文件放在一个文件夹中,同时确保文件夹中没有无关文件，并输入此文件夹路径')
        else:
            print('输入错误')
            file_name = ''
            continue
        check_select = input('请输入要检查表的类型：1：住房公积金信息 2：养老保险信息 3：水费缴纳信息 4：电费缴纳信息 5：燃气费缴纳信息 6：公共信用综合评价的企业登记注册信息 '
                             '7：企业填报信息 8：融资统计信息地方平台维度和金融机构维度 9：入驻金融机构信息 10：获贷企业信息 11：民营企业统计信息 12：职工医疗保险费基本信息 '
                             '13：被执行人信息 14：裁判文书终审判决信息 15：企业资质证书信息 16：企业荣誉表彰信息 17：职工医疗保险费欠缴信息 18：职工医疗保险费变更信息 '
                             '19：新型农业经营主体保险信息 20：新型农业经营主体补贴信息 21：种植大户清单信息 22：新型农业经营主体基本信息')
        print('开始检查')
        try:
            if check_select == '8':
                double_check(file_type, file_name, fill, check_date)
            else:
                main_check(file_type, file_name, check_select, fill, check_date)
            print('检查完成')
        except Exception as e:
            print('检查失败')
            print(e)
        # if check_select == '8':
        #     double_check(file_type, file_name, fill, check_date)
        # else:
        #     main_check(file_type, file_name, check_select, fill, check_date)
        # print('检查完成')

    # check_date = str_to_date('202410')
    # fill = PatternFill("solid", fgColor='FFFF00')
    # house_check(1, '（改）住房公积金信息.xlsm', fill, check_date)
